"""
Tests for creep_model.io.parser.ExcelCreepParser.

Uses two data sources:
  1. A minimal SYNTHETIC workbook built with openpyxl (via the
     `synthetic_workbook` fixture below), for exact, hand-checkable
     assertions about parsing logic in isolation.
  2. The REAL data/raw/CreepData.xlsx (via `real_data_path` from
     tests/conftest.py), for a smoke-test that the parser still works
     end-to-end against actual data -- skipped if the file isn't present
     (data/ is git-ignored, see .gitignore).
"""
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from creep_model.io.parser import ExcelCreepParser


@pytest.fixture
def synthetic_workbook(tmp_path) -> Path:
    """
    Builds a minimal 2-test-sheet workbook (+ Home, MetaData) matching the
    real CreepData.xlsx layout:
        - Home: present, must be dropped and never treated as a test.
        - MetaData: RunCode, StressPa, AgeDays, nRun, PrintQuality, Failure.
        - One row per test sheet: [Time_s, Extension_mm, TempTime_mins,
          Temperature_degC], with strain/time and temp/time on DIFFERENT
          (temp coarser) row counts, and a NaN-padded tail on the temp
          columns once they run out -- matching the real workbook's shape.
    """
    path = tmp_path / "SyntheticCreepData.xlsx"
    wb = openpyxl.Workbook()

    # Default sheet becomes "Home"
    home = wb.active
    home.title = "Home"

    # MetaData sheet
    meta = wb.create_sheet("MetaData")
    meta.append(["RunCode", "StressPa", "AgeDays", "nRun", "PrintQuality", "Failure"])
    meta.append(["T.1", 1.0e7, 2, 1, "High", 0])
    meta.append(["T.2", 2.0e7, 3, 2, "Standard", 0])
    # A row referencing a sheet that doesn't exist yet -- parser should skip
    # it with a warning, not crash (mirrors the real "upcoming test slot"
    # workflow noted in io/parser.py's comments).
    meta.append(["T.3.NOT_YET_RUN", 3.0e7, 4, 3, "High", 0])

    # Test sheet 1: pre-load row (dropped) + 4 real extension readings;
    # temperature sampled on a coarser, independent time base.
    t1 = wb.create_sheet("T.1")
    t1.append(["Time_s", "Extension_mm", "TempTime_mins", "Temperature_degC"])
    t1.append([0, 0.00, 0.0, 20.0])    # pre-load row -- dropped by parser
    t1.append([10, 0.10, 5.0, 21.0])
    t1.append([20, 0.20, 10.0, 22.0])
    t1.append([30, 0.30, None, None])  # strain continues, temp already exhausted
    t1.append([40, 0.40, None, None])

    # Test sheet 2: simpler, no NaN padding needed.
    t2 = wb.create_sheet("T.2")
    t2.append(["Time_s", "Extension_mm", "TempTime_mins", "Temperature_degC"])
    t2.append([0, 0.00, 0.0, 19.0])   # pre-load row -- dropped
    t2.append([10, 0.50, 10.0, 19.5])
    t2.append([20, 1.00, 20.0, 20.0])

    wb.save(path)
    return path


class TestExcelCreepParserInit:
    def test_raises_file_not_found_for_missing_path(self, tmp_path):
        missing = tmp_path / "does_not_exist.xlsx"
        with pytest.raises(FileNotFoundError, match="Cannot find data file"):
            ExcelCreepParser(missing)

    def test_accepts_string_path(self, synthetic_workbook):
        # Should not raise even though a str, not a Path, is passed.
        parser = ExcelCreepParser(str(synthetic_workbook))
        assert parser.filepath == synthetic_workbook


class TestLoadExperimentSynthetic:
    def test_home_sheet_is_dropped(self, synthetic_workbook):
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        assert "Home" not in experiment.tests

    def test_loads_expected_test_sheets(self, synthetic_workbook):
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        assert set(experiment.tests.keys()) == {"T.1", "T.2"}

    def test_missing_sheet_is_skipped_not_raised(self, synthetic_workbook, capsys):
        """MetaData references T.3.NOT_YET_RUN, which has no matching sheet
        -- parser.py deliberately prints a warning and continues, rather
        than raising, to support adding metadata rows for not-yet-run
        tests."""
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        assert "T.3.NOT_YET_RUN" not in experiment.tests
        captured = capsys.readouterr()
        assert "T.3.NOT_YET_RUN" in captured.out
        assert "Skipping" in captured.out

    def test_preload_row_is_excluded(self, synthetic_workbook):
        """First extension reading (the pre-load row, t=0) must be dropped
        -- time_series/strain_series should start from the SECOND row."""
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        test = experiment.tests["T.1"]
        assert test.time_series[0] == pytest.approx(10.0)
        np.testing.assert_allclose(test.time_series, [10.0, 20.0, 30.0, 40.0])

    def test_strain_computed_from_extension_and_gauge_length(self, synthetic_workbook):
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        test = experiment.tests["T.2"]
        # gauge_length_mm = 20.0 (config.py); extension 0.5, 1.0 -> strain 0.025, 0.05
        np.testing.assert_allclose(test.strain_series, [0.025, 0.05])

    def test_temperature_series_shorter_than_strain_series(self, synthetic_workbook):
        """Confirms the parser preserves the real-world shape where
        temperature is recorded far less often than strain -- temp rows
        with NaN are dropped via df_temp.dropna(), so temperature_readings
        ends up SHORTER than time_series/strain_series, on its own
        independent time base (temp_time_series).

        NOTE: unlike the extension/strain data, df_temp is NOT subject to
        the pre-load-row drop (that `.iloc[1:]` only applies to df_ext) --
        so temperature_readings here includes the t=0 reading too, giving
        3 valid temp rows against 4 strain rows (not 2)."""
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        test = experiment.tests["T.1"]
        assert len(test.strain_series) == 4
        assert len(test.temperature_readings) == 3  # 3 non-NaN temp rows (incl. t=0)
        np.testing.assert_allclose(test.temperature_readings, [20.0, 21.0, 22.0])

    def test_temp_time_series_converted_minutes_to_seconds(self, synthetic_workbook):
        """parser.py multiplies TempTime_mins by 60 to get seconds, to put
        it on the same units as Time_s (even though it's a different,
        coarser grid). Includes the t=0 reading since df_temp isn't
        subject to the pre-load-row drop (see note above)."""
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        test = experiment.tests["T.1"]
        np.testing.assert_allclose(test.temp_time_series, [0.0, 5.0 * 60, 10.0 * 60])

    def test_metadata_injected_correctly(self, synthetic_workbook):
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        test = experiment.tests["T.1"]
        assert test.applied_stress_MPa == pytest.approx(10.0)  # 1.0e7 Pa -> 10 MPa
        assert test.age_days == 2
        assert test.print_quality == "High"

        test2 = experiment.tests["T.2"]
        assert test2.applied_stress_MPa == pytest.approx(20.0)  # 2.0e7 Pa -> 20 MPa
        assert test2.print_quality == "Standard"

    def test_test_id_matches_sheet_name(self, synthetic_workbook):
        experiment = ExcelCreepParser(synthetic_workbook).load_experiment()
        assert experiment.tests["T.1"].test_id == "T.1"
        assert experiment.tests["T.2"].test_id == "T.2"


class TestLoadExperimentRealData:
    """Smoke tests against the actual CreepData.xlsx -- skip if not present."""

    @pytest.fixture(autouse=True)
    def _load(self, real_data_path: Path):
        if not real_data_path.exists():
            pytest.skip(f"Real data file not found at {real_data_path} (data/ is git-ignored).")
        self.experiment = ExcelCreepParser(real_data_path).load_experiment()

    def test_loads_24_test_sheets(self):
        assert len(self.experiment.tests) == 24

    def test_no_home_or_metadata_sheet_leaks_into_tests(self):
        assert "Home" not in self.experiment.tests
        assert "MetaData" not in self.experiment.tests

    def test_all_tests_have_shorter_or_equal_temperature_series(self):
        """Sanity check on the real data's known shape (per project notes:
        strain ~480 points vs temperature ~24 points per test)."""
        for test in self.experiment.tests.values():
            if test.is_empty:
                continue
            assert len(test.temperature_readings) <= len(test.time_series)

    def test_known_failed_specimens_present(self):
        assert "S.30.2" in self.experiment.tests
        assert "H.30.3" in self.experiment.tests